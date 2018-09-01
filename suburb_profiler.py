from openpyxl import load_workbook

#Variables
postcodes=[3087,3141,3178,3175,3138,3977]
sheets=['G 08', 'G 29','G 53']
totalcols=[7,4,11]
firstrows = [11,11,14]
lastrows = [42,29,33]

catrow=[0,0,12];#category row
totalrows=[0,0,35]
firstcols=[0,0,2]
lastcols=[0,0,10]

#Fixed parameters
prepend='GCP_POA'
ext='.xlsx'

#files=[prepend+str(postcode)+ext for postcode in postcodes]
#print(files)
num_pcodes=len(postcodes)
num_sheets=len(sheets)
value_sheets=[]
for postcode in postcodes:
    filename=prepend+str(postcode)+ext
    wb = load_workbook(filename=filename, read_only=True)
    value_lists=[]
    for i in range(num_sheets):
        ws = wb[sheets[i]]
        if sheets[i] == 'G 53':
            value_list = [ws.cell(row=totalrows[i],column=c).value for c in range(firstcols[i],lastcols[i]+1)]
        else:
            value_list = [ws.cell(row=r,column=totalcols[i]).value for r in range(firstrows[i],lastrows[i]+1)]
        value_list.insert(0,postcode)
        value_lists.append(value_list)
    value_sheets.append(value_lists)

title_list=[]
for i in range(num_sheets):
    ws = wb[sheets[i]]
    title_list.append(ws.cell(row=4,column=1).value)

category_lists=[]
for i in range(num_sheets):
    ws = wb[sheets[i]]
    if sheets[i] == 'G 53':
        category_list = [ws.cell(row=catrow[i],column=c).value for c in range(firstcols[i],lastcols[i]+1)]
    else:
        category_list = [ws.cell(row=r,column=1).value for r in range(firstrows[i],lastrows[i]+1)]
    category_list.insert(0,'PostCode')
    category_lists.append(category_list)

#print(category_lists)
#print(value_sheets)

numrows=[lastrows[i]-firstrows[i] for i in range(num_sheets)]
numcols=[lastcols[i]-firstcols[i] for i in range(num_sheets)]
sum_sheets=[]
for p in range(num_pcodes):
    sum_lists=[]
    for i in range(num_sheets):
        total=0
        if sheets[i] == 'G 53':
            for c in range(1,numcols[i]+2):
                total = total + value_sheets[p][i][c]
        else:
            for r in range(1,numrows[i]+2):
                total = total + value_sheets[p][i][r]
        #sum(value_sheets[p][i])
        sum_lists.append(total)
    sum_sheets.append(sum_lists)

percentage_sheets=[]
for p in range(num_pcodes):
    percentage_lists=[]
    for i in range(num_sheets):
        '''percentage_list=[]
        for r in range(1,numrows[i]+2):
            fval=float(value_sheets[p][i][r])
            fsum=float(sum_sheets[p][i])
            percentage_list.append(fval*100.0/fsum)'''
        if sheets[i] == 'G 53':
            percentage_list = [value_sheets[p][i][c]*100/sum_sheets[p][i] for c in range(1,numcols[i]+2)]
        else:
            percentage_list = [value_sheets[p][i][r]*100/sum_sheets[p][i] for r in range(1,numrows[i]+2)]
        percentage_list.insert(0,postcodes[p])
        percentage_lists.append(percentage_list)
    percentage_sheets.append(percentage_lists)

#print(sum_sheets)

from openpyxl import Workbook
wb = Workbook()



'''for i in range(num_sheets):
    wsw = wb.create_sheet(category=sheets[i])
    wsw.append(category_lists[i])
    for c in range(num_pcodes):
        wsw.append(value_sheets[c][i])'''

for i in range(num_sheets):
    wsw = wb.create_sheet(title=sheets[i])
    wsw.cell(row=1,column=1).value=title_list[i]

    if sheets[i] == 'G 53':
        for r in range(firstcols[i],lastcols[i]+2):
            wsw.cell(row=r-firstcols[i]+2,column=1).value=category_lists[i][r-firstcols[i]]
    else:
        for r in range(firstrows[i],lastrows[i]+2):
            wsw.cell(row=r-firstrows[i]+2,column=1).value=category_lists[i][r-firstrows[i]]

percent_col_offset = num_pcodes+3
for c in range(num_pcodes):
    #print(postcodes[c])
    for i in range(num_sheets):
        wsw = wb[sheets[i]]
        if sheets[i] == 'G 53':
            for r in range(firstcols[i],lastcols[i]+2):
                wsw.cell(row=r-firstcols[i]+2,column=c+2).value=value_sheets[c][i][r-firstcols[i]]
                wsw.cell(row=r-firstcols[i]+2,column=percent_col_offset+c+2).value=percentage_sheets[c][i][r-firstcols[i]]
        else:
            for r in range(firstrows[i],lastrows[i]+2):
                wsw.cell(row=r-firstrows[i]+2,column=c+2).value=value_sheets[c][i][r-firstrows[i]]
                wsw.cell(row=r-firstrows[i]+2,column=percent_col_offset+c+2).value=percentage_sheets[c][i][r-firstrows[i]]

#save the file
wb.save('compilation.xlsx')
